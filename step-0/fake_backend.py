from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME
import random
from datetime import datetime, timedelta
import numpy as np

num_rows = 999

wb = Workbook()
ws = wb.active

ws.merge_cells('A1:C1')
ws['A1'] = "Reporting - Players"
ws['A1'].alignment = Alignment(horizontal='left')

ws.merge_cells('A2:C2')
ws['A2'] = "Run time: 10.03.2024\n01:02:03+00:00 UTC"
ws['A2'].alignment = Alignment(horizontal='left', wrap_text=True)

headers = ["First Deposit Date", "Playercode", "Total deposits"]
ws.append(headers)

used_codes = set()

for i in range(num_rows):
    # below line is to ensure that majority of players are always in the current month
    current_month_probability = 0.8
    if random.random() < current_month_probability:
        today = datetime.now()
        start_of_month = today.replace(day=1)
        days_in_month = (today - start_of_month).days
        random_day = random.randint(0, days_in_month)
        date_value = start_of_month + timedelta(days=random_day)
    else:
        # below line is to generate dates from the last year to make it at least work somehow
        date_value = datetime.now() - timedelta(days=random.randint(1, 365))
    
    player_code = str(np.random.randint(12345678, 12349999))
    while player_code in used_codes:
        player_code = str(np.random.randint(12345678, 12349999))
    used_codes.add(player_code)
    
    total_deposit = random.choice([10, 20, round(random.uniform(10, 2000), 2)])
    
    ws.append([date_value, player_code, total_deposit])
    if date_value:
        ws.cell(row=ws.max_row, column=1).number_format = FORMAT_DATE_DATETIME

output_excel_file = '../step-2/fake_backend_data.xlsx'
wb.save(filename=output_excel_file)

print(f"Excel file saved as {output_excel_file}")
